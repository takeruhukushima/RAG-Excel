import os
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill
import random
from datetime import datetime, timedelta


def create_sales_data_excel(output_dir: str):
    """売上データのExcelファイルを作成"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "2024年売上"
    
    # ヘッダー
    headers = ["日付", "商品名", "カテゴリ", "数量", "単価", "売上金額", "担当者"]
    ws.append(headers)
    
    # ヘッダーのスタイル
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    # データ
    products = [
        ("ノートパソコン", "電子機器", 150000),
        ("マウス", "周辺機器", 2500),
        ("キーボード", "周辺機器", 8000),
        ("モニター", "電子機器", 35000),
        ("USBメモリ", "記憶装置", 1500),
        ("外付けHDD", "記憶装置", 12000),
        ("プリンター", "OA機器", 25000),
        ("スキャナー", "OA機器", 18000),
    ]
    
    staff = ["佐藤", "鈴木", "高橋", "田中", "伊藤"]
    
    start_date = datetime(2024, 1, 1)
    
    for i in range(100):
        date = start_date + timedelta(days=random.randint(0, 364))
        product, category, price = random.choice(products)
        quantity = random.randint(1, 10)
        amount = price * quantity
        person = random.choice(staff)
        
        ws.append([
            date.strftime("%Y-%m-%d"),
            product,
            category,
            quantity,
            price,
            amount,
            person
        ])
    
    # 2023年売上シート
    ws2 = wb.create_sheet("2023年売上")
    ws2.append(headers)
    
    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    start_date_2023 = datetime(2023, 1, 1)
    for i in range(80):
        date = start_date_2023 + timedelta(days=random.randint(0, 364))
        product, category, price = random.choice(products)
        quantity = random.randint(1, 8)
        amount = price * quantity
        person = random.choice(staff)
        
        ws2.append([
            date.strftime("%Y-%m-%d"),
            product,
            category,
            quantity,
            price,
            amount,
            person
        ])
    
    filepath = Path(output_dir) / "売上データ.xlsx"
    wb.save(filepath)
    print(f"✓ 作成: {filepath}")


def create_employee_data_excel(output_dir: str):
    """社員データのExcelファイルを作成"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "社員一覧"
    
    # ヘッダー
    headers = ["社員ID", "氏名", "部署", "役職", "入社日", "年齢", "給与"]
    ws.append(headers)
    
    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    # データ
    departments = ["営業部", "開発部", "総務部", "人事部", "経理部"]
    positions = ["部長", "課長", "主任", "一般"]
    names = ["佐藤太郎", "鈴木花子", "高橋一郎", "田中美咲", "伊藤健太", 
             "渡辺久美", "山本隆", "中村さくら", "小林誠", "加藤恵"]
    
    for i in range(50):
        emp_id = f"EMP{i+1:04d}"
        name = random.choice(names) if i >= len(names) else names[i % len(names)]
        dept = random.choice(departments)
        position = random.choice(positions)
        hire_date = datetime(random.randint(2010, 2023), random.randint(1, 12), random.randint(1, 28))
        age = random.randint(25, 60)
        salary = random.randint(300, 800) * 10000
        
        ws.append([emp_id, name, dept, position, hire_date.strftime("%Y-%m-%d"), age, salary])
    
    # スキル一覧シート
    ws2 = wb.create_sheet("スキル管理")
    skill_headers = ["社員ID", "氏名", "Python", "Java", "JavaScript", "SQL", "Excel"]
    ws2.append(skill_headers)
    
    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    for i in range(30):
        emp_id = f"EMP{i+1:04d}"
        name = random.choice(names)
        skills = [random.choice(["初級", "中級", "上級", "エキスパート", "-"]) for _ in range(5)]
        ws2.append([emp_id, name] + skills)
    
    filepath = Path(output_dir) / "社員データ.xlsx"
    wb.save(filepath)
    print(f"✓ 作成: {filepath}")


def create_inventory_data_excel(output_dir: str):
    """在庫データのExcelファイルを作成"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "在庫一覧"
    
    # ヘッダー
    headers = ["商品コード", "商品名", "カテゴリ", "在庫数", "最低在庫数", "単価", "仕入先"]
    ws.append(headers)
    
    header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    header_font = Font(bold=True, color="000000")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    # データ
    products = [
        ("デスクトップPC", "電子機器", "A社"),
        ("ノートPC", "電子機器", "B社"),
        ("タブレット", "電子機器", "C社"),
        ("スマートフォン", "電子機器", "D社"),
        ("マウス", "周辺機器", "E社"),
        ("キーボード", "周辺機器", "E社"),
        ("ヘッドセット", "周辺機器", "F社"),
        ("Webカメラ", "周辺機器", "G社"),
        ("USBケーブル", "アクセサリ", "H社"),
        ("LANケーブル", "アクセサリ", "H社"),
    ]
    
    for i in range(50):
        code = f"PRD{i+1:05d}"
        product, category, supplier = random.choice(products)
        stock = random.randint(0, 500)
        min_stock = random.randint(10, 50)
        price = random.randint(500, 200000)
        
        ws.append([code, product, category, stock, min_stock, price, supplier])
    
    # 発注履歴シート
    ws2 = wb.create_sheet("発注履歴")
    order_headers = ["発注日", "商品コード", "商品名", "発注数", "単価", "合計金額", "納期"]
    ws2.append(order_headers)
    
    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    start_date = datetime(2024, 1, 1)
    for i in range(60):
        order_date = start_date + timedelta(days=random.randint(0, 300))
        code = f"PRD{random.randint(1, 50):05d}"
        product, _, _ = random.choice(products)
        quantity = random.randint(10, 100)
        price = random.randint(500, 200000)
        total = quantity * price
        delivery_date = order_date + timedelta(days=random.randint(7, 30))
        
        ws2.append([
            order_date.strftime("%Y-%m-%d"),
            code,
            product,
            quantity,
            price,
            total,
            delivery_date.strftime("%Y-%m-%d")
        ])
    
    filepath = Path(output_dir) / "在庫データ.xlsx"
    wb.save(filepath)
    print(f"✓ 作成: {filepath}")


def create_project_data_excel(output_dir: str):
    """プロジェクトデータのExcelファイルを作成"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "プロジェクト一覧"
    
    # ヘッダー
    headers = ["プロジェクトID", "プロジェクト名", "顧客名", "開始日", "終了予定日", "進捗率", "予算", "担当PM"]
    ws.append(headers)
    
    header_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    # データ
    project_names = [
        "ECサイトリニューアル", "社内システム刷新", "モバイルアプリ開発",
        "データ分析基盤構築", "AIチャットボット導入", "クラウド移行",
        "セキュリティ強化", "業務効率化ツール開発", "顧客管理システム"
    ]
    
    clients = ["A株式会社", "B商事", "C工業", "D銀行", "Eホールディングス"]
    pms = ["山田太郎", "鈴木花子", "佐藤一郎", "田中美咲"]
    
    start_date = datetime(2023, 1, 1)
    
    for i in range(30):
        proj_id = f"PROJ{i+1:04d}"
        proj_name = random.choice(project_names)
        client = random.choice(clients)
        start = start_date + timedelta(days=random.randint(0, 600))
        duration = random.randint(30, 365)
        end = start + timedelta(days=duration)
        progress = random.randint(0, 100)
        budget = random.randint(100, 5000) * 10000
        pm = random.choice(pms)
        
        ws.append([
            proj_id,
            proj_name,
            client,
            start.strftime("%Y-%m-%d"),
            end.strftime("%Y-%m-%d"),
            f"{progress}%",
            budget,
            pm
        ])
    
    filepath = Path(output_dir) / "プロジェクトデータ.xlsx"
    wb.save(filepath)
    print(f"✓ 作成: {filepath}")

# ================== メイン実行 ==================
def main():
    """メイン処理"""
    global _in_memory_index
    
    print("=== Excel RAG System ===\n")
    
    # 1. Tantivyインデックスをメモリ上に構築
    print("Building in-memory Tantivy index...")
    _in_memory_index = build_tantivy_index(EXCEL_DIRECTORY)
    print("Index built successfully!\n")
    
    # 2. LangGraphの構築
    app = create_graph()
    
    # 3. ユーザーからの質問
    query = input("質問を入力してください: ")
    
    # 4. RAGの実行
    
    # ★★★ ここを修正 ★★★
    # SystemMessage と HumanMessage を1つの HumanMessage に統合する
    # -----------------------------------------------------------------
    system_prompt = "あなたはExcelファイルの内容を検索して回答するアシスタントです。search_excel_filesツールで関連ファイルを検索し、retrieve_from_excelツールで詳細情報を取得してください。"
    
    combined_prompt = f"""{system_prompt}

---

質問: {query}
"""
    
    initial_state = {
        "messages": [
            # [SystemMessage, HumanMessage] の組み合わせが 
            # 'contents is not specified' エラーを引き起こしているため、
            # 1つの HumanMessage に統合して問題を回避する
            HumanMessage(content=combined_prompt)
        ],
        "excel_files": [],
        "retrieved_context": "",
        "final_answer": ""
    }
    # -----------------------------------------------------------------
    
    print("\n処理中...\n")
    
    # グラフの最終結果（ENDノード）は、AgentState全体を返す
    result = app.invoke(initial_state)
    
    # 5. 結果の表示
    print("\n=== 回答 ===")
    
    # generate_final_answerノード（または私の前回の修正）で 
    # "final_answer" キーに回答が格納されているはず
    if "final_answer" in result and result["final_answer"]:
        print(result["final_answer"])
    else:
        # フォールバック：最後のメッセージを表示
        print("デバッグ: 最終回答キーが見つかりません。最後のメッセージを表示します。")
        if "messages" in result and result["messages"]:
            print(result["messages"][-1].content)
        else:
            print("エラー: 回答を生成できませんでした。")


if __name__ == "__main__":
    main()