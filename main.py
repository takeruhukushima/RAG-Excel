import os
import json
import re  # ← ★ これを追加
from pathlib import Path
from typing import List, Dict, Any, TypedDict

import tantivy
import openpyxl
import numpy as np
from dotenv import load_dotenv
from langchain_google_genai import ChatGoogleGenerativeAI, GoogleGenerativeAIEmbeddings
from langgraph.graph import StateGraph, END
from langgraph.prebuilt import ToolNode
# ★ ToolMessage をインポート
from langchain_core.messages import HumanMessage, AIMessage, SystemMessage, ToolMessage
from langchain_core.tools import tool

# .envファイルから環境変数を読み込み
load_dotenv()

# ================== 設定 ==================
GEMINI_API_KEY = os.getenv("GEMINI_API_KEY")
if not GEMINI_API_KEY:
    raise ValueError("GEMINI_API_KEY が .env ファイルに設定されていません")

EXCEL_DIRECTORY = os.getenv("EXCEL_DIRECTORY", "./excel_data")
CHUNK_SIZE = int(os.getenv("CHUNK_SIZE", "500"))


# ================== グローバルなインメモリインデックス ==================
_in_memory_index = None

def build_tantivy_index(directory: str) -> tantivy.Index:
    """Excelファイルのメタデータと内容をメモリ上でインデックス化"""
    
    # インデックスのスキーマ定義
    schema_builder = tantivy.SchemaBuilder()
    schema_builder.add_text_field("path", stored=True)
    schema_builder.add_text_field("filename", stored=True)
    schema_builder.add_text_field("content", stored=True)
    schema_builder.add_text_field("sheet_name", stored=True)
    schema = schema_builder.build()
    
    # インメモリインデックスの作成
    index = tantivy.Index(schema)
    
    writer = index.writer()
    
    # ディレクトリ内のExcelファイルを走査
    for excel_file in Path(directory).rglob("*.xlsx"):
        if excel_file.name.startswith("~$"):  # 一時ファイルをスキップ
            continue
            
        try:
            wb = openpyxl.load_workbook(excel_file, data_only=True)
            
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                
                # シートの内容をテキスト化
                content_parts = []
                for row in sheet.iter_rows(values_only=True):
                    row_text = " ".join([str(cell) for cell in row if cell is not None])
                    if row_text.strip():
                        content_parts.append(row_text)
                
                content = "\n".join(content_parts)
                
                # ドキュメントを追加
                doc = tantivy.Document()
                doc.add_text("path", str(excel_file))
                doc.add_text("filename", excel_file.name)
                doc.add_text("content", content)
                doc.add_text("sheet_name", sheet_name)
                writer.add_document(doc)
                
            wb.close()
            print(f"Indexed: {excel_file.name}")
            
        except Exception as e:
            print(f"Error indexing {excel_file}: {e}")
    
    writer.commit()
    return index

# ↓↓↓ main.py の 100行目あたりにあるこの関数を丸ごと置き換え ↓↓↓

def search_tantivy_index(index: tantivy.Index, query: str, limit: int = 5) -> List[Dict[str, str]]:
    """Tantivyインデックスを検索"""
    searcher = index.searcher()
    
    # ★★★ ここから修正 ★★★
    
    # paperqa のコード (529行目) が示すように、
    # QueryParser を手動で作成するのではなく、
    # index オブジェクトの .parse_query() メソッドを直接呼び出すのが正しい方法です。
    
    # 修正前 (すべての間違った試み):
    # query_parser = tantivy.QueryParser.for_index(...)
    # query_parser = index.query_parser(...)
    # query_parser = tantivy.query.QueryParser.for_index(...)

    # 修正後 (paperqa と同じ方法):
    query_fields = ["content", "filename"]
    
    try:
        parsed_query = index.parse_query(query, query_fields)
    except ValueError as e:
        # クエリが空だったり、構文が不正だと tantivy がエラーを出す
        print(f"Tantivy query parse error: {e}. Returning empty results.")
        return [] # エラー時は空のリストを返す
        
    # ★★★ ここまで修正 ★★★
    
    results = searcher.search(parsed_query, limit)
    
    docs = []
    for score, doc_address in results.hits:
        doc = searcher.doc(doc_address)
        docs.append({
            "path": doc["path"][0],
            "filename": doc["filename"][0],
            "sheet_name": doc["sheet_name"][0],
            "content": doc["content"][0][:500],  # プレビュー用
            "score": score
        })
    
    return docs

# ↑↑↑ ここまで置き換え ↑↑↑
# # ================== Excelのチャンク化とベクトル化 ==================
def chunk_excel_content(excel_path: str, chunk_size: int = CHUNK_SIZE) -> List[Dict[str, Any]]:
    """Excelファイルをチャンクに分割"""
    chunks = []
    
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        
        content_parts = []
        for row in sheet.iter_rows(values_only=True):
            row_text = " ".join([str(cell) for cell in row if cell is not None])
            if row_text.strip():
                content_parts.append(row_text)
        
        full_content = "\n".join(content_parts)
        
        # チャンクに分割
        for i in range(0, len(full_content), chunk_size):
            chunk_text = full_content[i:i+chunk_size]
            chunks.append({
                "text": chunk_text,
                "metadata": {
                    "source": excel_path,
                    "sheet": sheet_name,
                    "chunk_id": i // chunk_size
                }
            })
    
    wb.close()
    return chunks


def retrieve_relevant_chunks(chunks: List[Dict], query: str, embeddings_model, top_k: int = 3) -> List[Dict]:
    """ベクトル類似度でチャンクを検索"""
    # クエリの埋め込み
    query_embedding = embeddings_model.embed_query(query)
    
    # チャンクの埋め込み
    chunk_texts = [chunk["text"] for chunk in chunks]
    chunk_embeddings = embeddings_model.embed_documents(chunk_texts)
    
    # コサイン類似度を計算
    similarities = []
    for chunk_emb in chunk_embeddings:
        similarity = np.dot(query_embedding, chunk_emb) / (
            np.linalg.norm(query_embedding) * np.linalg.norm(chunk_emb)
        )
        similarities.append(similarity)
    
    # Top-kを取得
    top_indices = np.argsort(similarities)[-top_k:][::-1]
    
    relevant_chunks = [chunks[i] for i in top_indices]
    return relevant_chunks


# ================== LangGraph用のツール定義 ==================
@tool
def search_excel_files(query: str) -> str:
    """
    Tantivyを使ってExcelファイルを検索し、関連するExcelファイルのパスを返す。
    
    Args:
        query: 検索クエリ
    
    Returns:
        検索結果のJSON文字列
    """
    global _in_memory_index
    if _in_memory_index is None:
        raise ValueError("インデックスが初期化されていません")
    
    results = search_tantivy_index(_in_memory_index, query, limit=3)
    return json.dumps(results, ensure_ascii=False, indent=2)


@tool
def retrieve_from_excel(excel_path: str, query: str) -> str:
    """
    指定されたExcelファイルからクエリに関連する情報を取得する。
    
    Args:
        excel_path: Excelファイルのパス
        query: 検索クエリ
    
    Returns:
        関連するチャンクのテキスト
    """
    embeddings = GoogleGenerativeAIEmbeddings(
        model="models/text-embedding-004",
        google_api_key=GEMINI_API_KEY
    )
    
    chunks = chunk_excel_content(excel_path)
    relevant_chunks = retrieve_relevant_chunks(chunks, query, embeddings, top_k=3)
    
    result_text = "\n\n---\n\n".join([
        f"[{chunk['metadata']['sheet']}]\n{chunk['text']}"
        for chunk in relevant_chunks
    ])
    
    return result_text


# ================== LangGraphのステート定義 ==================
class AgentState(TypedDict):
    messages: List[Any]
    excel_files: List[Dict]
    retrieved_context: str
    final_answer: str
    original_question: str


# ★★★ ここからが修正の核心 ★★★

# ツールを定義
tools = [search_excel_files, retrieve_from_excel]

# 1. ツールがバインドされていない「素の」モデルを定義
model = ChatGoogleGenerativeAI(
    model="gemini-2.5-flash",
    google_api_key=GEMINI_API_KEY,
    temperature=0.7
)

# 2. ツールがバインドされたモデルを定義
model_with_tools = model.bind_tools(tools)


# ================== LangGraphのノード定義 ==================
def should_continue(state: AgentState) -> str:
    """次のステップを決定"""
    last_message = state["messages"][-1]
    
    # ツール呼び出しがある場合
    if hasattr(last_message, "tool_calls") and last_message.tool_calls:
        return "tools"
    
    return "end"

# ↓↓↓ main.py の 253行目あたりにあるこの関数を丸ごと置き換え ↓↓↓

def call_model(state: AgentState):
    """
    LLMを呼び出す（バグ回避ロジック入り）
    """
    last_message = state["messages"][-1]
    
    if isinstance(last_message, ToolMessage):
        # ツール実行後:
        # "contents is not specified" バグを回避するため、
        # メッセージ履歴をクリーンな単一プロンプトに再構築する
        
        print("--- [Debug] Re-constructing prompt to continue agent loop...")
        
        # 1. ツール結果を取得
        tool_result = last_message.content
        
        # 2. ★★★ 修正点 ★★★
        # 履歴を検索するのではなく、stateから元の質問を直接取得
        original_query = state.get("original_question", "（元の質問が見つかりません）")

        # 3. 新しいプロンプトを構築
        new_prompt = f"""
あなたはExcelファイルについて回答するエージェントです。
あなたは今、ステップの途中にいます。

[元の質問]
{original_query}

[直前に実行したツール（search_excel_files または retrieve_from_excel）の結果]
{tool_result}

[あなたの次のタスク]
上記の「元の質問」と「ツールの結果」を考慮してください。
- もし「ツールの結果」だけで「元の質問」に完全に回答できるなら、最終回答を生成してください。
- もし、ファイルの中身をさらに詳しく調べる必要があるなら（例: `search_excel_files`の結果が返ってきた場合）、「retrieve_from_excel」ツールを呼び出してください。
- ツールを呼び出す際は、`excel_path`と`query`引数を正しく設定してください。

あなたの思考と次のアクション（最終回答 または ツール呼び出し）を生成してください。
"""
        
        # 履歴をクリーンな [HumanMessage] にして、
        # ツールがバインドされた `model_with_tools` を呼び出す
        response = model_with_tools.invoke([HumanMessage(content=new_prompt)])
        
    else:
        # ツール実行前 (最初の呼び出し)
        print("--- [Debug] Calling model WITH tools (deciding action)...")
        response = model_with_tools.invoke(state["messages"])
    
    return {"messages": [response]}

# ↑↑↑ ここまで置き換え ↑↑↑
# # # ↑↑↑ ここまで置き換え ↑↑↑
# ================== LangGraphの構築 ==================
def create_graph():
    """LangGraphを構築"""
    workflow = StateGraph(AgentState)
    
    # ノードを追加 (generate は削除)
    workflow.add_node("agent", call_model)
    workflow.add_node("tools", ToolNode(tools))
    
    # エッジを追加
    workflow.set_entry_point("agent")
    
    # "end" の行き先を "generate" から END に変更
    workflow.add_conditional_edges("agent", should_continue, {
        "tools": "tools",
        "end": END
    })
    workflow.add_edge("tools", "agent")
    
    return workflow.compile()

# ================== メイン実行 ==================
def main():
    """メイン処理"""
    global _in_memory_index
    
    print("=== Excel RAG System ===\n")
    
    # 1. Tantivyインデックスをメモリ上に構築
    print("Building in-memory Tantivy index...")
    _in_memory_index = build_tantivy_index(EXCEL_DIRECTORY)
    print("Index built successfully!\n")
    
    # 2. LangGraphの構築
    app = create_graph()
    
    # 3. ユーザーからの質問
    query = input("質問を入力してください: ")
    
    # 4. RAGの実行
    
    # SystemMessage と HumanMessage を1つの HumanMessage に統合
    system_prompt = "あなたはExcelファイルの内容を検索して回答するアシスタントです。search_excel_filesツールで関連ファイルを検索し、retrieve_from_excelツールで詳細情報を取得してください。"
    
    combined_prompt = f"""{system_prompt}

---

質問: {query}
"""
    
    initial_state = {
        "messages": [
            HumanMessage(content=combined_prompt)
        ],
        "excel_files": [],
        "retrieved_context": "",
        "final_answer": "",
        "original_question": query  # ← ★ これを追加
    }
    
    print("\n処理中...\n")
    
    # グラフの最終結果（ENDノード）は、AgentState全体を返す
    result = app.invoke(initial_state)
    
    # 5. 結果の表示
    print("\n=== 回答 ===")
    
    # 最終回答は "messages" の最後に入っている
    if "messages" in result and result["messages"]:
        last_message = result["messages"][-1]
        
        # Gemini の .content は文字列かリストの場合がある
        if isinstance(last_message.content, str):
            print(last_message.content)
        elif isinstance(last_message.content, list) and len(last_message.content) > 0:
            first_part = last_message.content[0]
            if isinstance(first_part, dict) and 'text' in first_part:
                print(first_part['text'])
            else:
                print(str(last_message.content))
        else:
             print("エラー: 予期しないAIメッセージ形式です。")
    else:
        print("エラー: 回答を生成できませんでした。")


if __name__ == "__main__":
    main()