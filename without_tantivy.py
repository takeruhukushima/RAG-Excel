import os
import json
import re
from pathlib import Path
from typing import List, Dict, Any, TypedDict
import time
import openpyxl
from dotenv import load_dotenv

from langchain_chroma import Chroma
from langchain_core.documents import Document 

from langchain_google_genai import ChatGoogleGenerativeAI, GoogleGenerativeAIEmbeddings
from langgraph.graph import StateGraph, END
from langgraph.prebuilt import ToolNode
from langchain_core.messages import HumanMessage, AIMessage, SystemMessage, ToolMessage
from langchain_core.tools import tool

load_dotenv()

# ================== 設定 ==================
GEMINI_API_KEY = os.getenv("GEMINI_API_KEY")
if not GEMINI_API_KEY:
    raise ValueError("GEMINI_API_KEY が .env ファイルに設定されていません")

EXCEL_DIRECTORY = os.getenv("EXCEL_DIRECTORY", "./excel_data")
CHUNK_SIZE = int(os.getenv("CHUNK_SIZE", "2000"))  # 500 -> 2000に増やす

# ================== グローバル変数 ==================
_embedding_model = None
_vector_store = None

def chunk_excel_content(excel_path: str, chunk_size: int = CHUNK_SIZE) -> List[Dict[str, Any]]:
    """
    Excelファイルをチャンクに分割
    with_tantivy.py と同じシンプルなロジック
    """
    chunks = []
    
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            
            # 全行をテキスト化
            content_parts = []
            for row in sheet.iter_rows(values_only=True):
                row_text = " ".join([str(cell) for cell in row if cell is not None])
                if row_text.strip():
                    content_parts.append(row_text)
            
            full_content = "\n".join(content_parts)
            
            # チャンクに分割
            for i in range(0, len(full_content), chunk_size):
                chunk_text = full_content[i:i+chunk_size]
                chunks.append({
                    "text": chunk_text,
                    "metadata": {
                        "source": str(excel_path),
                        "sheet": sheet_name,
                        "chunk_id": i // chunk_size
                    }
                })
        
        wb.close()
    except Exception as e:
        print(f"Error chunking {excel_path}: {e}")
        
    return chunks

def build_in_memory_vector_store(directory: str):
    """
    起動時にすべてのExcelをチャンク化し、ChromaのインメモリDBに格納する
    """
    global _embedding_model, _vector_store
    
    print("Building in-memory Chroma vector store...")
    
    _embedding_model = GoogleGenerativeAIEmbeddings(
        model="models/text-embedding-004",
        google_api_key=GEMINI_API_KEY
    )
    
    _vector_store = Chroma(
        collection_name="excel_data_in_memory",
        embedding_function=_embedding_model
    )
    
    all_chunks = []
    
    for excel_file in Path(directory).rglob("*.xlsx"):
        if excel_file.name.startswith("~$"):
            continue
        
        print(f"  Chunking: {excel_file.name}")
        chunks = chunk_excel_content(str(excel_file))
        all_chunks.extend(chunks)

    if not all_chunks:
        print("No chunks found. Vector store is empty.")
        return

    print(f"Converting {len(all_chunks)} chunks to Documents...")
    documents = []
    for chunk in all_chunks:
        doc = Document(
            page_content=chunk['text'],
            metadata=chunk['metadata'] 
        )
        documents.append(doc)

    print(f"Adding {len(documents)} documents to Chroma...")
    
    # 一度に全て追加（Embeddingはバッチ処理されるため高速）
    _vector_store.add_documents(documents)
    print(f"  Successfully added {len(documents)} documents!")
    
    print("In-memory Chroma vector store built successfully!")

# ================== ツール定義 ==================
@tool
def search_excel_context(query: str) -> str:
    """
    インメモリのChromaベクトルストアを検索し、クエリに関連するExcelのチャンクを返す。
    """
    global _vector_store

    if _vector_store is None:
        return "エラー: インメモリ・ベクトルストアが初期化されていません。"

    print(f"--- [Debug] Searching Chroma vector store for: {query}")

    # ★改善: より少ない件数で正確な結果のみを返す
    top_k = 3  # 10 -> 3に削減
    fetch_k = 10  # 50 -> 10に削減
    
    print(f"--- [Debug] Retrieving Top {top_k} MMR results (fetching {fetch_k})...")
    relevant_docs = _vector_store.max_marginal_relevance_search(query, k=top_k, fetch_k=fetch_k)
    
    print("--- [Debug] Retrieved Docs: ---")
    for i, doc in enumerate(relevant_docs):
        print(f"  [Doc {i+1}] Source: {doc.metadata.get('source', 'N/A')}, Sheet: {doc.metadata.get('sheet', 'N/A')}")
        # ★改善: 全内容を表示（切り詰めない）
        print(f"  Content:\n{doc.page_content}")
        print("  " + "="*50)
    print("---------------------------------")
    
    # ★重要: 結果をより明確に構造化し、行ごとに番号を付ける
    result_parts = []
    for i, doc in enumerate(relevant_docs, 1):
        source = doc.metadata.get('source', 'N/A').split('/')[-1]
        sheet = doc.metadata.get('sheet', 'N/A')
        
        # チャンク内容を行ごとに分割
        lines = [line.strip() for line in doc.page_content.split('\n') if line.strip()]
        
        # 各行に番号を付けて明示
        numbered_lines = '\n'.join([f"  {j+1}. {line}" for j, line in enumerate(lines)])
        
        result_parts.append(
            f"【データソース {i}】\n"
            f"ファイル: {source}\n"
            f"シート: {sheet}\n"
            f"内容（各行に番号付き）:\n{numbered_lines}\n"
        )
    
    result_text = "\n".join(result_parts)
    
    return result_text

# ================== ステート定義 ==================
class AgentState(TypedDict):
    messages: List[Any]
    original_question: str

tools = [search_excel_context]

model = ChatGoogleGenerativeAI(
    model="gemini-2.5-flash",
    google_api_key=GEMINI_API_KEY,
    temperature=0.7
)

model_with_tools = model.bind_tools(tools)

# ================== ノード定義 ==================
def should_continue(state: AgentState) -> str:
    """次のステップを決定"""
    last_message = state["messages"][-1]
    
    if hasattr(last_message, "tool_calls") and last_message.tool_calls:
        return "tools"
    
    return "end"

def call_model(state: AgentState):
    """LLMを呼び出す"""
    last_message = state["messages"][-1]
    
    if isinstance(last_message, ToolMessage):
        print("--- [Debug] Re-constructing prompt to generate final answer...")
        
        tool_result = last_message.content
        original_query = state.get("original_question", "（元の質問が見つかりません）")

        # ★★★ 最も厳格なプロンプト ★★★
        new_prompt = f"""
あなたは正確性を最優先するデータアナリストです。
以下の「検索結果」に**完全に一致する情報だけ**を使用して、「元の質問」に答えてください。

【絶対に守るルール】
1. 検索結果の「内容」セクションに**明記されている数字だけ**を使用すること
2. 「ノートパソコン」または「ノートPC」という商品名が含まれる行の「単価」の値だけを答えること
3. 行番号を参照して、どの行から取得した情報かを明示すること
4. 検索結果に存在しない数字は絶対に言わないこと
5. 不確実な場合は「分かりません」と答えること

[元の質問]
{original_query}

[検索結果]
{tool_result}

[回答フォーマット]
以下の形式で答えてください：
- データソース1、シート名、行番号X: 単価 XXXXX円
- データソース2、シート名、行番号Y: 単価 YYYYY円

検索結果に該当情報がない場合は「検索結果にノートパソコンの料金情報は見つかりませんでした」と答えてください。
"""
        
        response = model.invoke([HumanMessage(content=new_prompt)])
        
    else:
        print("--- [Debug] Calling model WITH tools (deciding action)...")
        response = model_with_tools.invoke(state["messages"])
    
    return {"messages": [response]}

# ================== グラフ構築 ==================
def create_graph():
    """LangGraphを構築"""
    workflow = StateGraph(AgentState)
    
    workflow.add_node("agent", call_model)
    workflow.add_node("tools", ToolNode(tools))
    
    workflow.set_entry_point("agent")
    
    workflow.add_conditional_edges("agent", should_continue, {
        "tools": "tools",
        "end": END
    })
    workflow.add_edge("tools", "agent")
    
    return workflow.compile()

# ================== メイン実行 ==================
def main():
    """メイン処理"""
    print("=== Excel RAG System (Chroma In-Memory Mode) ===\n")
    
    build_in_memory_vector_store(EXCEL_DIRECTORY)
    
    app = create_graph()
    
    query = input("\n質問を入力してください: ")
    
    # ★検索クエリの改善: 質問全体を使う
    system_prompt = f"""
あなたはExcelファイルについて回答するアシスタントです。
ユーザーの質問に答えるために、必ず `search_excel_context` ツールを呼び出してください。

ツールを呼び出す際、`query`引数には質問から重要なキーワードを抽出して設定してください。
例:
- 元の質問: 「ノートパソコンの料金を教えて」 -> `query`引数: "ノートパソコン 料金"
- 元の質問: 「佐藤さんの部署は?」 -> `query`引数: "佐藤 部署"
"""
    
    combined_prompt = f"""{system_prompt}

---

質問: {query}
"""
    
    initial_state = {
        "messages": [
            HumanMessage(content=combined_prompt)
        ],
        "original_question": query
    }
    
    print("\n処理中...\n")
    
    result = app.invoke(initial_state)
    
    print("\n=== 回答 ===")
    
    if "messages" in result and result["messages"]:
        last_message = result["messages"][-1]
        
        if isinstance(last_message.content, str):
            print(last_message.content)
        elif isinstance(last_message.content, list) and len(last_message.content) > 0:
            first_part = last_message.content[0]
            if isinstance(first_part, dict) and 'text' in first_part:
                print(first_part['text'])
            else:
                print(str(last_message.content))
        else:
             print("エラー: 予期しないAIメッセージ形式です。")
    else:
        print("エラー: 回答を生成できませんでした。")


if __name__ == "__main__":
    main()